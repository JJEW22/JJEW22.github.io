#!/usr/bin/env python3
"""
Convert March Madness bracket Excel files to JSON format.

This script reads Excel bracket files and exports them to a standardized JSON format
that can be used by both the Svelte frontend and Python win probability calculator.
"""

import os
import json
import argparse
import csv
from typing import Dict, List, Optional
from openpyxl import load_workbook


# Constants matching bracketStructure.js
REGION_ORDER = ['East', 'West', 'South', 'Midwest']

MATCHUP_PAIRS = [
    (1, 16), (8, 9), (5, 12), (4, 13),
    (6, 11), (3, 14), (7, 10), (2, 15)
]

# Excel cell mappings for reading brackets
CELL_MAPPINGS = {
    'East': {
        'side': 'left',
        'position': 'top',
        'r1_rows': [7, 11, 15, 19, 23, 27, 31, 35],
        'r2_rows': [8, 12, 16, 20, 24, 28, 32, 36],
        's16_rows': [10, 18, 26, 34],
        'e8_rows': [14, 30],
        'f4_row': 22,
        'seed_col': 'B',
        'team_col': 'C',
        'r2_col': 'E',
        's16_col': 'H',
        'e8_col': 'K',
        'f4_col': 'N',
        'r1_start_index': 0,
        # Score columns (to the right of team names, toward center)
        'r1_score_col': 'D',
        'r2_score_col': 'F',
        's16_score_col': 'I',
        'e8_score_col': 'L',
        'f4_score_col': 'O'
    },
    'West': {
        'side': 'left',
        'position': 'bottom',
        'r1_rows': [42, 46, 50, 54, 58, 62, 66, 70],
        'r2_rows': [43, 47, 51, 55, 59, 63, 67, 71],
        's16_rows': [45, 53, 61, 69],
        'e8_rows': [49, 65],
        'f4_row': 57,
        'seed_col': 'B',
        'team_col': 'C',
        'r2_col': 'E',
        's16_col': 'H',
        'e8_col': 'K',
        'f4_col': 'N',
        'r1_start_index': 8,
        # Score columns (to the right of team names, toward center)
        'r1_score_col': 'D',
        'r2_score_col': 'F',
        's16_score_col': 'I',
        'e8_score_col': 'L',
        'f4_score_col': 'O'
    },
    'South': {
        'side': 'right',
        'position': 'top',
        'r1_rows': [7, 11, 15, 19, 23, 27, 31, 35],
        'r2_rows': [8, 12, 16, 20, 24, 28, 32, 36],
        's16_rows': [10, 18, 26, 34],
        'e8_rows': [14, 30],
        'f4_row': 22,
        'seed_col': 'AM',
        'team_col': 'AL',
        'r2_col': 'AJ',
        's16_col': 'AG',
        'e8_col': 'AD',
        'f4_col': 'AA',
        'r1_start_index': 16,
        # Score columns (to the left of team names, toward center)
        'r1_score_col': 'AK',
        'r2_score_col': 'AI',
        's16_score_col': 'AF',
        'e8_score_col': 'AC',
        'f4_score_col': 'Z'
    },
    'Midwest': {
        'side': 'right',
        'position': 'bottom',
        'r1_rows': [42, 46, 50, 54, 58, 62, 66, 70],
        'r2_rows': [43, 47, 51, 55, 59, 63, 67, 71],
        's16_rows': [45, 53, 61, 69],
        'e8_rows': [49, 65],
        'f4_row': 57,
        'seed_col': 'AM',
        'team_col': 'AL',
        'r2_col': 'AJ',
        's16_col': 'AG',
        'e8_col': 'AD',
        'f4_col': 'AA',
        'r1_start_index': 24,
        # Score columns (to the left of team names, toward center)
        'r1_score_col': 'AK',
        'r2_score_col': 'AI',
        's16_score_col': 'AF',
        'e8_score_col': 'AC',
        'f4_score_col': 'Z'
    }
}

# Championship cells
CHAMPIONSHIP = {
    'left_f4_winner': ('O', 39),   # East/West winner
    'right_f4_winner': ('W', 39),  # South/Midwest winner
    'champion': ('R', 44),
    # Score columns
    'left_f4_score': ('P', 39),    # East/West winner score
    'right_f4_score': ('V', 39),   # South/Midwest winner score
    # Final Four scores (at F4 row positions)
    'east_f4_score': ('O', 22),
    'west_f4_score': ('O', 57),
    'south_f4_score': ('Z', 22),
    'midwest_f4_score': ('Z', 57)
}


def load_teams_from_csv(filepath: str) -> Dict[str, dict]:
    """Load teams from CSV file and return dict mapping name to team data."""
    teams = {}
    with open(filepath, 'r', encoding='utf-8-sig') as f:  # utf-8-sig handles BOM
        reader = csv.DictReader(f)
        for row in reader:
            # Handle different column name cases
            name = row.get('Team') or row.get('TEAM') or row.get('name')
            seed_str = row.get('Seed') or row.get('SEED') or row.get('seed', '0')
            region = row.get('Region') or row.get('REGION') or row.get('region', '')
            
            if name:
                seed = int(seed_str) if seed_str else 0
                teams[name] = {'name': name, 'seed': seed, 'region': region}
    return teams


def find_team(name: str, teams: Dict[str, dict]) -> Optional[dict]:
    """Find team by name with fuzzy matching."""
    if not name:
        return None
    
    name = str(name).strip()
    
    # Direct match
    if name in teams:
        return teams[name].copy()
    
    # Case-insensitive match
    name_lower = name.lower()
    for team_name, team_data in teams.items():
        if team_name.lower() == name_lower:
            return team_data.copy()
    
    # Partial match
    for team_name, team_data in teams.items():
        if name_lower in team_name.lower() or team_name.lower() in name_lower:
            return team_data.copy()
    
    # Return basic team if not found
    return {'name': name, 'seed': 0, 'region': ''}


def get_cell_value(sheet, col: str, row: int) -> Optional[str]:
    """Get cell value from Excel sheet."""
    cell = sheet[f'{col}{row}']
    value = cell.value
    
    if value is None:
        return None
    
    # Handle formula results
    if hasattr(value, 'result'):
        value = value.result
    
    if isinstance(value, str):
        return value.strip() if value.strip() else None
    
    return str(value).strip() if value else None


def get_score_value(sheet, col: str, row: int) -> Optional[int]:
    """Get numeric score value from Excel sheet."""
    cell = sheet[f'{col}{row}']
    value = cell.value
    
    if value is None:
        return None
    
    # Handle formula results
    if hasattr(value, 'result'):
        value = value.result
    
    if value is None or value == '':
        return None
    
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def validate_score_consistency(score1: Optional[int], score2: Optional[int], 
                                winner_name: str, team1_name: str, team2_name: str,
                                game_desc: str) -> Optional[str]:
    """
    Validate that the winner has the higher score.
    
    Returns:
        Error message if validation fails, None if valid or scores not present
    """
    if score1 is None or score2 is None:
        return None  # Can't validate without both scores
    
    # Determine which team should have won based on scores
    if score1 > score2:
        score_winner = team1_name
    elif score2 > score1:
        score_winner = team2_name
    else:
        return f"  {game_desc}: Tie score ({score1}-{score2}) but winner is {winner_name}"
    
    # Check if score winner matches declared winner
    if score_winner != winner_name:
        return f"  {game_desc}: Score shows {score_winner} won ({score1}-{score2}) but winner is {winner_name}"
    
    return None


def create_empty_bracket() -> dict:
    """Create an empty bracket structure."""
    return {
        'round1': [None] * 32,
        'round2': [None] * 16,
        'round3': [None] * 8,
        'round4': [None] * 4,
        'round5': [None] * 2,
        'round6': [None] * 1,
        'winner': None
    }


def initialize_bracket_with_teams(teams: Dict[str, dict]) -> dict:
    """Initialize bracket with teams in correct positions."""
    bracket = create_empty_bracket()
    
    # Group teams by region
    regions = {region: [] for region in REGION_ORDER}
    for team_data in teams.values():
        region = team_data.get('region')
        if region in regions:
            regions[region].append(team_data)
    
    # Sort each region by seed
    for region in regions:
        regions[region].sort(key=lambda t: t.get('seed', 99))
    
    # Populate round 1
    game_index = 0
    for region_name in REGION_ORDER:
        region_teams = regions[region_name]
        
        for seed1, seed2 in MATCHUP_PAIRS:
            team1 = next((t for t in region_teams if t.get('seed') == seed1), None)
            team2 = next((t for t in region_teams if t.get('seed') == seed2), None)
            
            bracket['round1'][game_index] = {
                'team1': team1.copy() if team1 else None,
                'team2': team2.copy() if team2 else None,
                'winner': None,
                'region': region_name,
                'gameId': f'r1-{game_index}'
            }
            game_index += 1
    
    # Initialize later rounds with empty games
    for round_num in range(2, 7):
        round_key = f'round{round_num}'
        num_games = len(bracket[round_key])
        
        for i in range(num_games):
            bracket[round_key][i] = {
                'team1': None,
                'team2': None,
                'winner': None,
                'gameId': f'r{round_num}-{i}'
            }
    
    return bracket


def extract_bracket_from_excel(filepath: str, teams: Dict[str, dict], extract_scores: bool = False) -> dict:
    """Extract bracket data from Excel file."""
    workbook = load_workbook(filepath, data_only=True)
    
    # Try to find the bracket sheet
    sheet = None
    for sheet_name in ['madness', 'Bracket', 'Sheet1']:
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            break
    
    if sheet is None:
        sheet = workbook.active
    
    # Initialize bracket with teams
    bracket = initialize_bracket_with_teams(teams)
    
    # Track missing scores and score inconsistencies for error reporting
    missing_scores = []
    score_errors = []
    
    # Extract winners from each region
    for region_name, config in CELL_MAPPINGS.items():
        region_missing, region_errors = extract_region_winners(sheet, bracket, region_name, config, teams, extract_scores)
        missing_scores.extend(region_missing)
        score_errors.extend(region_errors)
    
    # Extract Final Four and Championship
    ff_missing, ff_errors = extract_final_four(sheet, bracket, teams, extract_scores)
    missing_scores.extend(ff_missing)
    score_errors.extend(ff_errors)
    
    # If extracting scores, check for errors
    if extract_scores:
        errors = []
        if missing_scores:
            errors.append("Missing scores for completed games:\n" + "\n".join(missing_scores))
        if score_errors:
            errors.append("Score inconsistencies (winner doesn't have higher score):\n" + "\n".join(score_errors))
        
        if errors:
            raise ValueError("\n\n".join(errors))
    
    return bracket


def extract_region_winners(sheet, bracket: dict, region: str, config: dict, teams: Dict[str, dict], extract_scores: bool = False) -> tuple:
    """Extract winners and optionally scores for a region from Excel sheet.
    
    Returns:
        Tuple of (missing_scores_list, score_errors_list)
    """
    start_idx = config['r1_start_index']
    missing_scores = []
    score_errors = []
    
    # Get score column names (may not exist in older configs)
    r1_score_col = config.get('r1_score_col') if extract_scores else None
    r2_score_col = config.get('r2_score_col') if extract_scores else None
    s16_score_col = config.get('s16_score_col') if extract_scores else None
    e8_score_col = config.get('e8_score_col') if extract_scores else None
    
    # Round 1 winners (from R2 cells) and scores
    for i, row in enumerate(config['r2_rows']):
        game_idx = start_idx + i
        winner_name = get_cell_value(sheet, config['r2_col'], row)
        if winner_name:
            winner = find_team(winner_name, teams)
            if winner:
                bracket['round1'][game_idx]['winner'] = winner
                
                # Populate round 2
                r2_idx = game_idx // 2
                if game_idx % 2 == 0:
                    bracket['round2'][r2_idx]['team1'] = winner
                else:
                    bracket['round2'][r2_idx]['team2'] = winner
        
        # Extract Round 1 scores from team rows (only if extract_scores=True)
        if r1_score_col and bracket['round1'][game_idx].get('winner'):
            team1_row = config['r1_rows'][i]
            team2_row = config['r1_rows'][i] + 2
            score1 = get_score_value(sheet, r1_score_col, team1_row)
            score2 = get_score_value(sheet, r1_score_col, team2_row)
            
            if score1 is not None:
                bracket['round1'][game_idx]['score1'] = score1
            else:
                missing_scores.append(f"  {region} Round 1 Game {i+1}: Missing score1 at {r1_score_col}{team1_row}")
            
            if score2 is not None:
                bracket['round1'][game_idx]['score2'] = score2
            else:
                missing_scores.append(f"  {region} Round 1 Game {i+1}: Missing score2 at {r1_score_col}{team2_row}")
            
            # Validate score consistency
            if score1 is not None and score2 is not None:
                team1_name = bracket['round1'][game_idx]['team1']['name'] if bracket['round1'][game_idx].get('team1') else None
                team2_name = bracket['round1'][game_idx]['team2']['name'] if bracket['round1'][game_idx].get('team2') else None
                winner_name = bracket['round1'][game_idx]['winner']['name']
                error = validate_score_consistency(score1, score2, winner_name, team1_name, team2_name,
                                                   f"{region} Round 1 Game {i+1}")
                if error:
                    score_errors.append(error)
    
    # Sweet 16 (Round 2 winners -> Round 3) and scores
    r2_start = start_idx // 2
    for i, row in enumerate(config['s16_rows']):
        r2_idx = r2_start + i
        winner_name = get_cell_value(sheet, config['s16_col'], row)
        if winner_name:
            winner = find_team(winner_name, teams)
            if winner:
                bracket['round2'][r2_idx]['winner'] = winner
                
                # Populate round 3
                r3_idx = r2_idx // 2
                if r2_idx % 2 == 0:
                    bracket['round3'][r3_idx]['team1'] = winner
                else:
                    bracket['round3'][r3_idx]['team2'] = winner
        
        # Extract Round 2 scores
        if r2_score_col and bracket['round2'][r2_idx].get('winner'):
            team1_row = config['r2_rows'][i * 2]
            team2_row = config['r2_rows'][i * 2 + 1]
            score1 = get_score_value(sheet, r2_score_col, team1_row)
            score2 = get_score_value(sheet, r2_score_col, team2_row)
            
            if score1 is not None:
                bracket['round2'][r2_idx]['score1'] = score1
            else:
                missing_scores.append(f"  {region} Round 2 Game {i+1}: Missing score1 at {r2_score_col}{team1_row}")
            
            if score2 is not None:
                bracket['round2'][r2_idx]['score2'] = score2
            else:
                missing_scores.append(f"  {region} Round 2 Game {i+1}: Missing score2 at {r2_score_col}{team2_row}")
            
            # Validate score consistency
            if score1 is not None and score2 is not None:
                team1_name = bracket['round2'][r2_idx]['team1']['name'] if bracket['round2'][r2_idx].get('team1') else None
                team2_name = bracket['round2'][r2_idx]['team2']['name'] if bracket['round2'][r2_idx].get('team2') else None
                winner_name = bracket['round2'][r2_idx]['winner']['name']
                error = validate_score_consistency(score1, score2, winner_name, team1_name, team2_name,
                                                   f"{region} Round 2 Game {i+1}")
                if error:
                    score_errors.append(error)
    
    # Elite 8 (Round 3 winners -> Round 4) and scores
    r3_start = start_idx // 4
    for i, row in enumerate(config['e8_rows']):
        r3_idx = r3_start + i
        winner_name = get_cell_value(sheet, config['e8_col'], row)
        if winner_name:
            winner = find_team(winner_name, teams)
            if winner:
                bracket['round3'][r3_idx]['winner'] = winner
                
                # Populate round 4
                r4_idx = r3_idx // 2
                if r3_idx % 2 == 0:
                    bracket['round4'][r4_idx]['team1'] = winner
                else:
                    bracket['round4'][r4_idx]['team2'] = winner
        
        # Extract Sweet 16 scores
        if s16_score_col and bracket['round3'][r3_idx].get('winner'):
            team1_row = config['s16_rows'][i * 2]
            team2_row = config['s16_rows'][i * 2 + 1]
            score1 = get_score_value(sheet, s16_score_col, team1_row)
            score2 = get_score_value(sheet, s16_score_col, team2_row)
            
            if score1 is not None:
                bracket['round3'][r3_idx]['score1'] = score1
            else:
                missing_scores.append(f"  {region} Sweet 16 Game {i+1}: Missing score1 at {s16_score_col}{team1_row}")
            
            if score2 is not None:
                bracket['round3'][r3_idx]['score2'] = score2
            else:
                missing_scores.append(f"  {region} Sweet 16 Game {i+1}: Missing score2 at {s16_score_col}{team2_row}")
            
            # Validate score consistency
            if score1 is not None and score2 is not None:
                team1_name = bracket['round3'][r3_idx]['team1']['name'] if bracket['round3'][r3_idx].get('team1') else None
                team2_name = bracket['round3'][r3_idx]['team2']['name'] if bracket['round3'][r3_idx].get('team2') else None
                winner_name = bracket['round3'][r3_idx]['winner']['name']
                error = validate_score_consistency(score1, score2, winner_name, team1_name, team2_name,
                                                   f"{region} Sweet 16 Game {i+1}")
                if error:
                    score_errors.append(error)
    
    # Final Four (Round 4 winner) and Elite 8 scores
    r4_idx = start_idx // 8  # 0=East, 1=West, 2=South, 3=Midwest
    winner_name = get_cell_value(sheet, config['f4_col'], config['f4_row'])
    if winner_name:
        winner = find_team(winner_name, teams)
        if winner:
            bracket['round4'][r4_idx]['winner'] = winner
            
            # Populate round 5 (Final Four)
            r5_idx = r4_idx // 2
            if r4_idx % 2 == 0:
                bracket['round5'][r5_idx]['team1'] = winner
            else:
                bracket['round5'][r5_idx]['team2'] = winner
    
    # Extract Elite 8 scores
    e8_score_col = config.get('e8_score_col') if extract_scores else None
    if e8_score_col and bracket['round4'][r4_idx].get('winner') and len(config['e8_rows']) >= 2:
        score1 = get_score_value(sheet, e8_score_col, config['e8_rows'][0])
        score2 = get_score_value(sheet, e8_score_col, config['e8_rows'][1])
        
        if score1 is not None:
            bracket['round4'][r4_idx]['score1'] = score1
        else:
            missing_scores.append(f"  {region} Elite 8: Missing score1 at {e8_score_col}{config['e8_rows'][0]}")
        
        if score2 is not None:
            bracket['round4'][r4_idx]['score2'] = score2
        else:
            missing_scores.append(f"  {region} Elite 8: Missing score2 at {e8_score_col}{config['e8_rows'][1]}")
        
        # Validate score consistency
        if score1 is not None and score2 is not None:
            team1_name = bracket['round4'][r4_idx]['team1']['name'] if bracket['round4'][r4_idx].get('team1') else None
            team2_name = bracket['round4'][r4_idx]['team2']['name'] if bracket['round4'][r4_idx].get('team2') else None
            winner_name = bracket['round4'][r4_idx]['winner']['name']
            error = validate_score_consistency(score1, score2, winner_name, team1_name, team2_name,
                                               f"{region} Elite 8")
            if error:
                score_errors.append(error)
    
    return missing_scores, score_errors


def extract_final_four(sheet, bracket: dict, teams: Dict[str, dict], extract_scores: bool = False) -> tuple:
    """Extract Final Four and Championship results with optional scores.
    
    Returns:
        Tuple of (missing_scores_list, score_errors_list)
    """
    missing_scores = []
    score_errors = []
    
    # Left semifinal winner (East/West)
    left_winner_name = get_cell_value(sheet, *CHAMPIONSHIP['left_f4_winner'])
    if left_winner_name:
        winner = find_team(left_winner_name, teams)
        if winner:
            bracket['round5'][0]['winner'] = winner
            bracket['round6'][0]['team1'] = winner
    
    # Right semifinal winner (South/Midwest)
    right_winner_name = get_cell_value(sheet, *CHAMPIONSHIP['right_f4_winner'])
    if right_winner_name:
        winner = find_team(right_winner_name, teams)
        if winner:
            bracket['round5'][1]['winner'] = winner
            bracket['round6'][0]['team2'] = winner
    
    # Champion
    champion_name = get_cell_value(sheet, *CHAMPIONSHIP['champion'])
    if champion_name:
        champion = find_team(champion_name, teams)
        if champion:
            bracket['round6'][0]['winner'] = champion
            bracket['winner'] = champion
    
    # Only extract scores if flag is set
    if not extract_scores:
        return missing_scores, score_errors
    
    # Extract Final Four scores (round 5) - only if game is completed
    # Left semifinal: East (O22) vs West (O57)
    if bracket['round5'][0].get('winner'):
        east_score = get_score_value(sheet, *CHAMPIONSHIP['east_f4_score'])
        west_score = get_score_value(sheet, *CHAMPIONSHIP['west_f4_score'])
        
        if east_score is not None:
            bracket['round5'][0]['score1'] = east_score
        else:
            col, row = CHAMPIONSHIP['east_f4_score']
            missing_scores.append(f"  Final Four Left Semifinal: Missing East score at {col}{row}")
        
        if west_score is not None:
            bracket['round5'][0]['score2'] = west_score
        else:
            col, row = CHAMPIONSHIP['west_f4_score']
            missing_scores.append(f"  Final Four Left Semifinal: Missing West score at {col}{row}")
        
        # Validate score consistency
        if east_score is not None and west_score is not None:
            team1_name = bracket['round5'][0]['team1']['name'] if bracket['round5'][0].get('team1') else None
            team2_name = bracket['round5'][0]['team2']['name'] if bracket['round5'][0].get('team2') else None
            winner_name = bracket['round5'][0]['winner']['name']
            error = validate_score_consistency(east_score, west_score, winner_name, team1_name, team2_name,
                                               "Final Four Left Semifinal")
            if error:
                score_errors.append(error)
    
    # Right semifinal: South (Z22) vs Midwest (Z57)
    if bracket['round5'][1].get('winner'):
        south_score = get_score_value(sheet, *CHAMPIONSHIP['south_f4_score'])
        midwest_score = get_score_value(sheet, *CHAMPIONSHIP['midwest_f4_score'])
        
        if south_score is not None:
            bracket['round5'][1]['score1'] = south_score
        else:
            col, row = CHAMPIONSHIP['south_f4_score']
            missing_scores.append(f"  Final Four Right Semifinal: Missing South score at {col}{row}")
        
        if midwest_score is not None:
            bracket['round5'][1]['score2'] = midwest_score
        else:
            col, row = CHAMPIONSHIP['midwest_f4_score']
            missing_scores.append(f"  Final Four Right Semifinal: Missing Midwest score at {col}{row}")
        
        # Validate score consistency
        if south_score is not None and midwest_score is not None:
            team1_name = bracket['round5'][1]['team1']['name'] if bracket['round5'][1].get('team1') else None
            team2_name = bracket['round5'][1]['team2']['name'] if bracket['round5'][1].get('team2') else None
            winner_name = bracket['round5'][1]['winner']['name']
            error = validate_score_consistency(south_score, midwest_score, winner_name, team1_name, team2_name,
                                               "Final Four Right Semifinal")
            if error:
                score_errors.append(error)
    
    # Extract Championship game scores (round 6) - only if game is completed
    if bracket['round6'][0].get('winner'):
        left_champ_score = get_score_value(sheet, *CHAMPIONSHIP['left_f4_score'])
        right_champ_score = get_score_value(sheet, *CHAMPIONSHIP['right_f4_score'])
        
        if left_champ_score is not None:
            bracket['round6'][0]['score1'] = left_champ_score
        else:
            col, row = CHAMPIONSHIP['left_f4_score']
            missing_scores.append(f"  Championship: Missing Team 1 score at {col}{row}")
        
        if right_champ_score is not None:
            bracket['round6'][0]['score2'] = right_champ_score
        else:
            col, row = CHAMPIONSHIP['right_f4_score']
            missing_scores.append(f"  Championship: Missing Team 2 score at {col}{row}")
        
        # Validate score consistency
        if left_champ_score is not None and right_champ_score is not None:
            team1_name = bracket['round6'][0]['team1']['name'] if bracket['round6'][0].get('team1') else None
            team2_name = bracket['round6'][0]['team2']['name'] if bracket['round6'][0].get('team2') else None
            winner_name = bracket['round6'][0]['winner']['name']
            error = validate_score_consistency(left_champ_score, right_champ_score, winner_name, team1_name, team2_name,
                                               "Championship")
            if error:
                score_errors.append(error)
    
    return missing_scores, score_errors


def convert_bracket_to_json(bracket: dict) -> dict:
    """Convert bracket to clean JSON-serializable format."""
    def clean_team(team):
        if not team:
            return None
        return {
            'name': team.get('name'),
            'seed': team.get('seed'),
            'region': team.get('region', '')
        }
    
    def clean_game(game):
        if not game:
            return None
        result = {
            'team1': clean_team(game.get('team1')),
            'team2': clean_team(game.get('team2')),
            'winner': clean_team(game.get('winner')),
            'gameId': game.get('gameId', ''),
            'region': game.get('region', '')
        }
        # Include scores if present
        if game.get('score1') is not None:
            result['score1'] = game['score1']
        if game.get('score2') is not None:
            result['score2'] = game['score2']
        return result
    
    return {
        'round1': [clean_game(g) for g in bracket['round1']],
        'round2': [clean_game(g) for g in bracket['round2']],
        'round3': [clean_game(g) for g in bracket['round3']],
        'round4': [clean_game(g) for g in bracket['round4']],
        'round5': [clean_game(g) for g in bracket['round5']],
        'round6': [clean_game(g) for g in bracket['round6']],
        'winner': clean_team(bracket.get('winner'))
    }


def convert_excel_to_json(excel_path: str, teams_path: str, output_path: str, extract_scores: bool = False):
    """Convert a single Excel bracket to JSON."""
    # Load teams
    teams = load_teams_from_csv(teams_path)
    
    # Extract bracket from Excel
    bracket = extract_bracket_from_excel(excel_path, teams, extract_scores=extract_scores)
    
    # Convert to clean JSON format
    json_bracket = convert_bracket_to_json(bracket)
    
    # Save to JSON
    with open(output_path, 'w') as f:
        json.dump(json_bracket, f, indent=2)
    
    print(f"Converted: {excel_path} -> {output_path}")


def convert_directory(input_dir: str, output_dir: str, teams_path: str, pattern: str = '*.xlsx', extract_scores: bool = False):
    """Convert all Excel brackets in a directory to JSON."""
    import glob
    
    os.makedirs(output_dir, exist_ok=True)
    
    excel_files = glob.glob(os.path.join(input_dir, pattern))
    print('found files', excel_files)
    if not excel_files:
        print(f"No Excel files found matching {pattern} in {input_dir}")
        return
    
    # Load teams once
    print('teams path', teams_path)
    teams = load_teams_from_csv(teams_path)
    
    for excel_path in excel_files:
        filename = os.path.basename(excel_path)
        # Change extension to .json
        json_filename = os.path.splitext(filename)[0] + '.json'
        output_path = os.path.join(output_dir, json_filename)
        
        try:
            bracket = extract_bracket_from_excel(excel_path, teams, extract_scores=extract_scores)
            json_bracket = convert_bracket_to_json(bracket)
            
            with open(output_path, 'w') as f:
                json.dump(json_bracket, f, indent=2)
            
            print(f"✓ {filename} -> {json_filename}")
        except Exception as e:
            print(f"✗ {filename}: {e}")


def main():
    parser = argparse.ArgumentParser(description='Convert Excel brackets to JSON format')
    parser.add_argument('--input', '-i', required=True, 
                       help='Input Excel file or directory containing Excel files')
    parser.add_argument('--output', '-o', required=True,
                       help='Output JSON file or directory')
    parser.add_argument('--teams', '-t', required=True,
                       help='Path to teams CSV file')
    parser.add_argument('--pattern', '-p', default='*.xlsx',
                       help='File pattern when converting directory (default: *.xlsx)')
    parser.add_argument('--score', '-s', action='store_true',
                       help='Extract game scores (for results bracket only). Errors if scores are missing for completed games.')
    
    args = parser.parse_args()
    
    if os.path.isdir(args.input):
        # Convert entire directory
        convert_directory(args.input, args.output, args.teams, args.pattern, extract_scores=args.score)
    else:
        # Convert single file
        convert_excel_to_json(args.input, args.teams, args.output, extract_scores=args.score)


if __name__ == '__main__':
    main()