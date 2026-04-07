#!/usr/bin/env python3
"""
Process pizza bracket ratings from xlsx into bracket results JSON.

Reads a spreadsheet where:
- Row 1: subheaders (division + round)
- Row 2: team name headers
- Row 3+: voter ratings (1-5 scale)

Determines winners using:
- 2-team matchups: each voter's higher rating = 1 vote. Tiebreaker: higher average.
- 3-team matchups: eliminate lowest average, then use 2-team voting on remaining two.

Usage:
    python processPizzaRatings.py --ratings pizzaBracketRatings.xlsx --bracket pizzaBracket.json --output pizzaBracketResults.json
"""

import json
import argparse
import os
import sys
from collections import defaultdict

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)


def load_ratings(xlsx_path, col_map):
    """Load voter ratings from xlsx, grouped by match."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    
    # Read all voter rows (starting from row 3)
    match_ratings = defaultdict(lambda: defaultdict(list))  # matchId -> teamName -> [ratings]
    
    voter_count = 0
    for row_idx in range(3, ws.max_row + 1):
        voter_name = ws.cell(row=row_idx, column=1).value
        if not voter_name or str(voter_name).strip() == '':
            continue
        
        has_any_rating = False
        for col_info in col_map:
            cell_val = ws.cell(row=row_idx, column=col_info['column']).value
            if cell_val is not None and cell_val != '':
                try:
                    rating = float(cell_val)
                    match_ratings[col_info['matchId']][col_info['teamName']].append(rating)
                    has_any_rating = True
                except (ValueError, TypeError):
                    pass
        
        if has_any_rating:
            voter_count += 1
    
    print(f"Loaded ratings from {voter_count} voters")
    return dict(match_ratings)


def determine_winner_2team(team1_name, team2_name, ratings1, ratings2):
    """
    Determine winner between 2 teams.
    Each voter's higher rating = 1 vote. Tiebreaker: higher average.
    Returns (winner_name, team1_votes, team2_votes, team1_avg, team2_avg)
    """
    if not ratings1 and not ratings2:
        return None, 0, 0, 0, 0
    
    votes1 = 0
    votes2 = 0
    
    # Pair up ratings (zip by index — same voter)
    min_len = min(len(ratings1), len(ratings2))
    for i in range(min_len):
        if ratings1[i] > ratings2[i]:
            votes1 += 1
        elif ratings2[i] > ratings1[i]:
            votes2 += 1
        # Equal = no vote for either
    
    avg1 = sum(ratings1) / len(ratings1) if ratings1 else 0
    avg2 = sum(ratings2) / len(ratings2) if ratings2 else 0
    
    if votes1 > votes2:
        winner = team1_name
    elif votes2 > votes1:
        winner = team2_name
    else:
        # Tiebreaker: higher average
        if avg1 > avg2:
            winner = team1_name
        elif avg2 > avg1:
            winner = team2_name
        else:
            winner = team1_name  # True tie — first team wins
            print(f"  ⚠️ Perfect tie between {team1_name} and {team2_name}, defaulting to {team1_name}")
    
    return winner, votes1, votes2, avg1, avg2


def determine_winner_3team(teams_ratings):
    """
    Determine winner among 3 teams.
    1. Eliminate lowest average
    2. Use 2-team voting between remaining two
    Returns (winner_name, results_dict) where results_dict has per-team votes/avg
    """
    # Calculate averages
    averages = {}
    for team_name, ratings in teams_ratings.items():
        if ratings:
            averages[team_name] = sum(ratings) / len(ratings)
        else:
            averages[team_name] = 0
    
    if len(averages) < 3:
        print(f"  ⚠️ Only {len(averages)} teams have ratings in triple match")
        # Fall back to whatever we have
        if len(averages) == 2:
            names = list(averages.keys())
            winner, v1, v2, a1, a2 = determine_winner_2team(
                names[0], names[1],
                teams_ratings[names[0]], teams_ratings[names[1]]
            )
            return winner, {
                names[0]: {'votes': v1, 'rating': a1},
                names[1]: {'votes': v2, 'rating': a2}
            }
        elif len(averages) == 1:
            name = list(averages.keys())[0]
            return name, {name: {'votes': 0, 'rating': averages[name]}}
        return None, {}
    
    # Step 1: Eliminate lowest average
    sorted_teams = sorted(averages.items(), key=lambda x: x[1])
    eliminated = sorted_teams[0][0]
    remaining = [t for t in averages.keys() if t != eliminated]
    
    print(f"  Triple: averages = {', '.join(f'{k}: {v:.2f}' for k, v in averages.items())}")
    print(f"  Eliminated (lowest avg): {eliminated} ({averages[eliminated]:.2f})")
    
    # Step 2: Vote between remaining two
    winner, v1, v2, a1, a2 = determine_winner_2team(
        remaining[0], remaining[1],
        teams_ratings[remaining[0]], teams_ratings[remaining[1]]
    )
    
    results = {
        remaining[0]: {'votes': v1, 'rating': a1},
        remaining[1]: {'votes': v2, 'rating': a2},
        eliminated: {'votes': None, 'rating': averages[eliminated]}
    }
    
    return winner, results


def process_bracket(bracket, match_ratings):
    """Process all matches and update bracket with results."""
    
    def process_match(match, ratings_for_match):
        """Process a single match and return updated match."""
        team_names = [t['name'] for t in match['teams'] if t['name'] != 'TBD' and t.get('source') != 'bye']
        
        if not ratings_for_match:
            print(f"  No ratings for match {match['id']} — skipping")
            return
        
        if match['type'] == 'triple':
            teams_ratings = {}
            for t in match['teams']:
                if t['name'] in ratings_for_match:
                    teams_ratings[t['name']] = ratings_for_match[t['name']]
            
            winner, results = determine_winner_3team(teams_ratings)
            
            if winner:
                match['winner'] = winner
                for team in match['teams']:
                    if team['name'] in results:
                        team['votes'] = results[team['name']]['votes']
                        team['rating'] = round(results[team['name']]['rating'], 2)
                
                print(f"  ✓ {match['id']}: {winner} wins!")
                for name, r in results.items():
                    v = f"{r['votes']}v" if r['votes'] is not None else "—"
                    print(f"    {name}: {v}, avg={r['rating']:.2f}")
        
        elif match['type'] == 'standard':
            # Get the two teams that have ratings
            rated_teams = [(t, ratings_for_match.get(t['name'], [])) for t in match['teams'] if t['name'] != 'TBD']
            
            if len(rated_teams) < 2:
                print(f"  Not enough rated teams for {match['id']}")
                return
            
            t1, r1 = rated_teams[0]
            t2, r2 = rated_teams[1]
            
            winner, v1, v2, a1, a2 = determine_winner_2team(t1['name'], t2['name'], r1, r2)
            
            if winner:
                match['winner'] = winner
                t1['votes'] = v1
                t1['rating'] = round(a1, 2)
                t2['votes'] = v2
                t2['rating'] = round(a2, 2)
                
                print(f"  ✓ {match['id']}: {winner} wins! ({v1}-{v2}, avg {a1:.2f} vs {a2:.2f})")
    
    # Process divisions round by round
    for div in bracket['divisions']:
        div_name = div['name']
        print(f"\n{'='*40}")
        print(f"  {div_name}")
        print(f"{'='*40}")
        
        for round_data in div['rounds']:
            print(f"\n  --- {round_data['name']} ---")
            for match in round_data['matches']:
                ratings = match_ratings.get(match['id'], {})
                process_match(match, ratings)
                
                # Propagate winner to next round
                if match['winner']:
                    propagate_winner(bracket, match)
    
    # Process finals
    print(f"\n{'='*40}")
    print(f"  FINALS")
    print(f"{'='*40}")
    for round_data in bracket['finals']['rounds']:
        print(f"\n  --- {round_data['name']} ---")
        for match in round_data['matches']:
            ratings = match_ratings.get(match['id'], {})
            process_match(match, ratings)


def propagate_winner(bracket, match):
    """Propagate a match winner to the next match that references this match's ID."""
    winner_name = match['winner']
    match_id = match['id']
    
    # Search all matches for one that has source = this match_id
    all_matches = []
    for div in bracket['divisions']:
        for round_data in div['rounds']:
            for m in round_data['matches']:
                all_matches.append(m)
    for round_data in bracket['finals']['rounds']:
        for m in round_data['matches']:
            all_matches.append(m)
    
    for m in all_matches:
        for team in m['teams']:
            if team.get('source') == match_id and (team['name'] == 'TBD' or team['name'] == ''):
                team['name'] = winner_name
                print(f"  → Propagated {winner_name} to {m['id']}")


def main():
    parser = argparse.ArgumentParser(description='Process pizza bracket ratings')
    parser.add_argument('--ratings', required=True, help='Path to ratings xlsx')
    parser.add_argument('--bracket', required=True, help='Path to bracket JSON')
    parser.add_argument('--column-map', help='Path to column map JSON (auto-generated with template)')
    parser.add_argument('--output', help='Output path (default: overwrite bracket)')
    
    args = parser.parse_args()
    
    # Load bracket
    with open(args.bracket, 'r') as f:
        bracket = json.load(f)
    
    # Load or generate column map
    if args.column_map and os.path.exists(args.column_map):
        with open(args.column_map) as f:
            col_map = json.load(f)
    else:
        # Auto-generate from bracket
        col_map = []
        col_idx = 2  # Start at column B (1-indexed)
        
        for div in bracket['divisions']:
            for round_data in div['rounds']:
                for match in round_data['matches']:
                    for ti, team in enumerate(match['teams']):
                        if team['name'] == 'TBD' or team.get('source') == 'bye':
                            continue
                        col_map.append({
                            'column': col_idx,
                            'matchId': match['id'],
                            'teamIndex': ti,
                            'teamName': team['name']
                        })
                        col_idx += 1
        
        for round_data in bracket['finals']['rounds']:
            for match in round_data['matches']:
                for ti, team in enumerate(match['teams']):
                    if team['name'] == 'TBD':
                        continue
                    col_map.append({
                        'column': col_idx,
                        'matchId': match['id'],
                        'teamIndex': ti,
                        'teamName': team['name']
                    })
                    col_idx += 1
    
    print(f"Column map: {len(col_map)} team columns")
    
    # Load ratings
    match_ratings = load_ratings(args.ratings, col_map)
    
    print(f"\nMatches with ratings: {len(match_ratings)}")
    for mid, teams in match_ratings.items():
        print(f"  {mid}: {', '.join(f'{t} ({len(r)} ratings)' for t, r in teams.items())}")
    
    # Process
    process_bracket(bracket, match_ratings)
    
    # Save
    output_path = args.output or args.bracket
    with open(output_path, 'w') as f:
        json.dump(bracket, f, indent=2)
    print(f"\nResults saved to {output_path}")


if __name__ == '__main__':
    main()